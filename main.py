import streamlit as st
import openpyxl
from dataclasses import dataclass
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import fitz
import tempfile
import os
from openpyxl import load_workbook
import pyexcel as p
from pyexcel import save_book_as
import mojimoji
import re
import pandas as pd
import win32com.client as win32
import datetime
import pythoncom
import shutil
import glob
import tempfile

# ===== アプリ専用 Temp（ここに一時ファイルを集約）=====
APP_TEMP_DIR = os.path.join(tempfile.gettempdir(), "excel_tenki_app")
os.makedirs(APP_TEMP_DIR, exist_ok=True)

TESSERACT_DIR_CANDIDATES = [
    r"C:\Program Files\Tesseract-OCR",
    r"C:\Program Files (x86)\Tesseract-OCR",
]
for _tesseract_dir in TESSERACT_DIR_CANDIDATES:
    _tesseract_exe = os.path.join(_tesseract_dir, "tesseract.exe")
    if os.path.exists(_tesseract_exe):
        if _tesseract_dir not in os.environ.get("PATH", ""):
            os.environ["PATH"] = _tesseract_dir + os.pathsep + os.environ.get("PATH", "")
        _tessdata_dir = os.path.join(_tesseract_dir, "tessdata")
        if os.path.isdir(_tessdata_dir):
            os.environ.setdefault("TESSDATA_PREFIX", _tessdata_dir)
        break

def cleanup_app_temp():
    """このアプリが作った一時ファイルだけ削除する（Temp全体は触らない）"""
    try:
        shutil.rmtree(APP_TEMP_DIR, ignore_errors=True)
        os.makedirs(APP_TEMP_DIR, exist_ok=True)
    except Exception:
        pass


SUPPLIER_TARGET_SHEET = {
    "オフィスインテリア": "見積明細（建築・内装）",
    "コマニー": "見積明細（コマニー）",
}


def validate_dest_cell(dest_cell: str, expected_col: str = "B") -> int:
    m = re.match(r"^([A-Z]+)(\d+)$", (dest_cell or "").upper().strip())
    if not m:
        raise ValueError("転記先セルは 'B3' の形式で入力してください")
    col, row = m.group(1), int(m.group(2))
    if col != expected_col:
        raise ValueError(f"転記先セルは {expected_col}列スタートで入力してください（例：{expected_col}{row}）")
    return row

def get_row_from_cell(cell_addr: str) -> int:
    m = re.match(r"^([A-Z]+)(\d+)$", (cell_addr or "").upper().strip())
    if not m:
        raise ValueError("セルは 'C2' の形式で入力してください")
    return int(m.group(2))


import shutil

def write_df_to_template_com(
    template_path: str,
    out_path: str,
    sheet_name: str,
    start_cell: str,
    df_src,
    col_map: dict,
):
    # ✅ 新規のときだけコピー（同一ファイルならコピーしない）
    if os.path.abspath(template_path) != os.path.abspath(out_path):
        shutil.copy(template_path, out_path)


    start_cell = (start_cell or "").strip().upper()
    start_row = validate_dest_cell(start_cell, expected_col="B")

    pythoncom.CoInitialize()
    excel = None
    wb = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        # UpdateLinks=0 でリンク更新を避ける（余計な警告・遅延防止）
        wb = excel.Workbooks.Open(out_path, UpdateLinks=0)
        ws = wb.Worksheets(sheet_name)

        n_rows = len(df_src)
        if n_rows == 0:
            raise ValueError("転記データが空です")

        # 1列ずつ書き込み（式列は触らない＝潰れない）
        for tmpl_col_letter, src_col_letter in col_map.items():
            c = column_index_from_string(tmpl_col_letter)  # テンプレ列番号

            # 縦1列の2次元配列（Excel Rangeに渡す形）
            col_values = []
            for i in range(n_rows):
                v = df_src.iloc[i][src_col_letter] if src_col_letter in df_src.columns else ""
                if v is None:
                    v = ""
                col_values.append([v])

            r0 = start_row
            r1 = start_row + n_rows - 1
            ws.Range(ws.Cells(r0, c), ws.Cells(r1, c)).Value = col_values

        # 数式再計算
        excel.CalculateFullRebuild()

        wb.Save()  # out_path に保存（ロゴ保持）
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except:
            pass
        pythoncom.CoUninitialize()




def get_col_map(company: str) -> dict:
    if company in ["オフィス（１シート）", "オフィス（複数シート）"]:
        return {
            "B": "C",  # 名称 ← 確定
            "C": "D",  # 商品記号
            "D": "E",  # 使用
            "E": "F",  # 数量
            "F": "G",  # 単位
            "K": "H",  # 単価
            
        }
    elif company == "帝国倉庫PDF":
        return {
            "B": "A",  # 品目
            "E": "D",  # 数量
            "F": "E",  # 単位
            "K": "F",  # 単価
            "I": "H",  # 備考
        }
    elif company == "コマニー":
        return {
            "B": "A",  # 名称
            "D": "C",  # 仕様・寸法 ← W寸法をここへ
            "E": "D",  # 数量
            "F": "E",  # 単位
            "K": "F",  # 単価
            "I": "H",  # 備考に案件番号を入れたいなら（必要なら追加）
        }

    else:
        raise ValueError(f"未対応の仕入先です: {company}")


@dataclass(frozen=True)
class Word:
    x0: float
    y0: float
    x1: float
    y1: float
    text: str


def _word_close(a: Word, b: Word, tolerance: float = 2.0) -> bool:
    return (
        abs(a.x0 - b.x0) <= tolerance
        and abs(a.y0 - b.y0) <= tolerance
        and abs(a.x1 - b.x1) <= tolerance
        and abs(a.y1 - b.y1) <= tolerance
    )


def _dedupe_words(words: list[Word]) -> list[Word]:
    deduped: list[Word] = []
    for w in sorted(words, key=lambda item: (round(item.y0, 1), item.x0, item.text)):
        if any(_word_close(w, existing) for existing in deduped):
            continue
        deduped.append(w)
    return deduped


def load_words(page: fitz.Page, y_min: float, y_max: float) -> list[Word]:
    words: list[Word] = []
    for x0, y0, x1, y1, text, *_rest in page.get_text("words"):
        if y_min <= y0 <= y_max:
            words.append(Word(float(x0), float(y0), float(x1), float(y1), str(text)))
    return words


def cluster_rows(words: list[Word], tolerance: float = 2.8) -> list[float]:
    ys = sorted({round(w.y0, 1) for w in words})
    rows: list[float] = []
    for y in ys:
        if not rows or abs(y - rows[-1]) > tolerance:
            rows.append(y)
    return rows


def row_for_y(y: float, rows: list[float]) -> float:
    return min(rows, key=lambda row_y: abs(row_y - y))


def build_detail_rows(page: fitz.Page) -> list[list[str]]:
    words = load_words(page, 205, 535)
    row_centers = [
        209.4,
        228.4,
        247.3,
        266.3,
        285.2,
        304.2,
        323.2,
        342.1,
        361.1,
        380.0,
        399.0,
        418.0,
        436.9,
        455.9,
        474.8,
        493.8,
        512.8,
        531.7,
    ]

    rows: dict[float, dict[int, list[tuple[float, str]]]] = {
        row: {i: [] for i in range(1, 8)} for row in row_centers
    }

    for w in words:
        row = row_for_y(w.y0, row_centers)
        if w.x0 < 90:
            col = 1
        elif w.x0 < 220:
            col = 2
        elif w.x0 < 300:
            col = 3
        elif w.x0 < 335:
            col = 4
        elif w.x0 < 390:
            col = 5
        elif w.x0 < 455:
            col = 6
        else:
            col = 7
        rows[row][col].append((w.x0, w.text))

    ordered_rows: list[list[str]] = [["大項目", "品目", "単価", "数量", "単位", "金額", "備考"]]
    for row_y in sorted(rows):
        cols = rows[row_y]
        values = []
        for idx in range(1, 8):
            parts = [text for _, text in sorted(cols[idx], key=lambda item: item[0])]
            values.append("".join(parts).strip())
        if any(values):
            # 備考欄の折り返しだけでできた行は、前の行の備考へ結合する。
            if (
                len(ordered_rows) > 1
                and not any(values[:6])
                and values[6]
            ):
                prev = ordered_rows[-1]
                prev[6] = (prev[6] + " " + values[6]).strip() if prev[6] else values[6]
                continue
            ordered_rows.append(values)
    return ordered_rows


def append_teisoh_misc_row(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    misc_row = pd.DataFrame(
        [["", "諸経費", 56000, 1, "式", 56000, ""]],
        columns=df.columns[:7],
    )
    return pd.concat([df, misc_row], ignore_index=True)


def normalize_teisoh_item_name(value: object) -> object:
    if not isinstance(value, str):
        return value

    text = value.strip()
    if text == "2t":
        return "2t車"
    if text == "4t":
        return "4t車"
    return text


def split_merged_quantity_and_price(quantity: object, unit_price: object) -> tuple[object, object]:
    def _split(value: object) -> tuple[str, str]:
        if not isinstance(value, str):
            return "", ""
        text = value.strip().replace(" ", "")
        m = re.match(r"^(-?\d{1,3})(\d{1,3}(?:,\d{3})+)$", text)
        if not m:
            return "", ""
        return m.group(1), m.group(2)

    quantity_text = "" if quantity is None else str(quantity).strip()
    unit_price_text = "" if unit_price is None else str(unit_price).strip()

    if quantity_text and not unit_price_text:
        q, p = _split(quantity_text)
        if q and p:
            return q, p

    if unit_price_text and not quantity_text:
        q, p = _split(unit_price_text)
        if q and p:
            return q, p

    return quantity, unit_price


def load_teisoh_detail_df(pdf_path: str) -> pd.DataFrame:
    doc = fitz.open(pdf_path)
    try:
        detail_rows: list[list[str]] = []
        if len(doc) > 1:
            for page_index in range(1, len(doc)):
                page_rows = build_detail_rows(doc[page_index])
                if not page_rows:
                    continue
                if not detail_rows:
                    detail_rows.extend(page_rows)
                else:
                    detail_rows.extend(page_rows[1:])
    finally:
        doc.close()

    source_df = pd.DataFrame(detail_rows[1:], columns=["A", "B", "C", "D", "E", "F", "G"]).fillna("")
    source_df["B"] = source_df["B"].map(normalize_teisoh_item_name)

    split_pairs = source_df.apply(
        lambda row: split_merged_quantity_and_price(row["D"], row["C"]),
        axis=1,
        result_type="expand",
    )
    source_df["D"] = split_pairs[0]
    source_df["C"] = split_pairs[1]

    output_rows: list[list[str]] = [["品目", "", "", "数量", "単位", "単価", "金額", "備考"]]
    for _, row in source_df.iterrows():
        item = str(row["B"]).strip()
        if not item:
            continue
        output_rows.append(
            [
                item,
                "",
                "",
                str(row["D"]).strip(),
                str(row["E"]).strip(),
                str(row["C"]).strip(),
                str(row["F"]).strip(),
                str(row["G"]).strip(),
            ]
        )

    output_rows.append(["諸経費", "", "", "1", "式", "56000", "56000", ""])
    df = pd.DataFrame(output_rows, columns=["A", "B", "C", "D", "E", "F", "G", "H"])
    df.index += 1
    return df



# ここから下が Streamlit のUI（関数の外）
st.set_page_config(page_title="Excel転記アプリ", layout="wide")


# ▼ 新規スタート（状態クリア）
if st.sidebar.button("🧹 新規スタート（データクリア）"):
    cleanup_app_temp()

    # ★ uploaderを作り直すための番号を進める
    st.session_state["uploader_reset_counter"] = st.session_state.get("uploader_reset_counter", 0) + 1

    for k in [
        "result_path",
        "template_path",
        "template_wb",
        "last_result_path",
        "download_name",
        "updated_template_df_map",
        "prev_mode",
        "original_template_path",
        "office_multi_start_sheet",
        "sheet_selector",
        "src_cell",
        "dest_cell",
        "output_name",
        "prev_company",
    ]:
        if k in st.session_state:
            del st.session_state[k]

    st.rerun()

# ===== 転記結果 自動復元 =====
RESULT_DIR = os.path.join(os.path.expanduser("~"), "Excel転記アプリ_results")
os.makedirs(RESULT_DIR, exist_ok=True)

def find_latest_result() -> str | None:
    files = glob.glob(os.path.join(RESULT_DIR, "*.xlsx"))
    if not files:
        return None
    return max(files, key=os.path.getmtime)

# ===== 転記ロック（同時実行防止）=====
LOCK_FILE = os.path.join(RESULT_DIR, "transfer.lock")
LOCK_TIMEOUT_SEC = 10 * 60  # 10分

def _now_ts():
    return datetime.datetime.now().timestamp()

def is_lock_stale(timeout_sec=LOCK_TIMEOUT_SEC) -> bool:
    try:
        with open(LOCK_FILE, "r", encoding="utf-8") as f:
            ts = float(f.read().strip() or "0")
        return (_now_ts() - ts) > timeout_sec
    except:
        return False

def acquire_lock() -> bool:
    # 古いロックは自動解除
    if os.path.exists(LOCK_FILE) and is_lock_stale():
        try:
            os.remove(LOCK_FILE)
        except:
            pass

    try:
        fd = os.open(LOCK_FILE, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(str(_now_ts()))
        return True
    except FileExistsError:
        return False

def release_lock():
    try:
        if os.path.exists(LOCK_FILE):
            os.remove(LOCK_FILE)
    except:
        pass


# --- SessionState 初期化（必須）---
if "last_result_path" not in st.session_state:
    st.session_state.last_result_path = None
if "template_wb" not in st.session_state:
    st.session_state.template_wb = None
if "template_path" not in st.session_state:
    st.session_state.template_path = None
if "original_template_path" not in st.session_state:
    st.session_state.original_template_path = None
if "updated_template_df_map" not in st.session_state:
    st.session_state.updated_template_df_map = {}

st.title("仕入先見積変換＆転記アプリ")


# --- 共通処理 ---
def save_uploaded_bytes(file_bytes, suffix=".xls"):
    os.makedirs(APP_TEMP_DIR, exist_ok=True)
    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix, dir=APP_TEMP_DIR)
    tmp_file.write(file_bytes)
    tmp_file.close()
    return tmp_file.name

def convert_xls_to_xlsx(xls_path):
    xlsx_path = xls_path + "x"
    p.save_book_as(file_name=xls_path, dest_file_name=xlsx_path)
    return xlsx_path

def get_data_until_blank(ws, start_col, start_row):
    data = []
    row = start_row
    while row <= ws.max_row:
        cell_val = ws.cell(row=row, column=start_col).value
        data.append((row, cell_val))
        row += 1
    return data
def normalize_sheet_name(name: str) -> str:
    # 全角数字を半角に寄せる（例：３→3）
    try:
        return mojimoji.zen_to_han(name)
    except Exception:
        return name


def consolidate_selected_sheets(wb, target_sheet_names, add_blank_row=True):
    """
    target_sheet_names の先頭シートに、2枚目以降のデータを追記して1枚にまとめる
    """
    if not target_sheet_names or len(target_sheet_names) <= 1:
        return

    base_ws = wb[target_sheet_names[0]]

    def is_row_blank(ws, r, max_col):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                return False
        return True

    base_max_col = base_ws.max_column
    base_last = base_ws.max_row
    while base_last > 1 and is_row_blank(base_ws, base_last, base_max_col):
        base_last -= 1
    write_row = base_last + 1

    for src_name in target_sheet_names[1:]:
        if src_name not in wb.sheetnames:
            continue

        src_ws = wb[src_name]
        max_col = max(base_max_col, src_ws.max_column)

        for r in range(1, src_ws.max_row + 1):
            row_vals = []
            all_blank = True

            for c in range(1, max_col + 1):
                v = src_ws.cell(row=r, column=c).value
                row_vals.append(v)
                if v is not None and str(v).strip() != "":
                    all_blank = False

            if all_blank:
                continue

            for c, v in enumerate(row_vals, start=1):
                base_ws.cell(row=write_row, column=c).value = v

            write_row += 1

        if add_blank_row:
            write_row += 1

        wb.remove(src_ws)


def office_int_multi(xlsx_path: str, start_sheet_name: str):
    """
    オフィスインテリア（複数シート版）
    ・プルダウンで選んだシート以降を処理
    ・不要行削除、文字整形、単価補完、列削除
    ・最後に対象シートを1枚に統合
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    all_sheet_names = wb.sheetnames

    if start_sheet_name not in all_sheet_names:
        wb.close()
        raise ValueError(f"指定した開始シートが見つかりません: {start_sheet_name}")

    start_idx = all_sheet_names.index(start_sheet_name)

    # ★ 選択したシート以降をすべて対象にする
    target_sheet_names = all_sheet_names[start_idx:]

    if not target_sheet_names:
        wb.close()
        raise ValueError("処理対象シートがありません")

    target_sheets = [wb[name] for name in target_sheet_names]

    for ws in target_sheets:
        # 0) 行削除
        for r in range(ws.max_row, 0, -1):
            b_val = get_cell_value_with_merge(ws, r, 2)
            b_norm = normalize_text(b_val)

            c_val = get_cell_value_with_merge(ws, r, 3)
            c_norm = normalize_text(c_val)

            cm_all_empty = True
            for col in range(3, 14):  # C～M
                v = get_cell_value_with_merge(ws, r, col)
                if normalize_text(v) != "":
                    cm_all_empty = False
                    break

            hi_text = ""
            for col in (8, 9, 10, 11):  # H, I, J, K
                hi_text += normalize_text(get_cell_value_with_merge(ws, r, col))

            if (
                ("件名" in b_norm)
                or ("名称" in b_norm)
                or ("合計" in c_norm)
                or ("内訳書" in hi_text or "内訳明細書" in hi_text)
                or cm_all_empty
            ):
                ws.delete_rows(r)

        # ① C列・I列の先頭スペース削除
        for row in range(1, ws.max_row + 1):
            v = get_cell_value_with_merge(ws, row, 3)  # C
            if isinstance(v, str):
                ws.cell(row=row, column=3).value = v.lstrip(" 　")

            v = get_cell_value_with_merge(ws, row, 9)  # I
            if isinstance(v, str):
                ws.cell(row=row, column=9).value = v.lstrip(" 　")

        # ② L列が空欄ならM列をコピー（ただしC列が小計は除外）
        for row in range(1, ws.max_row + 1):
            c_val = get_cell_value_with_merge(ws, row, 3)   # C
            l_val = get_cell_value_with_merge(ws, row, 12)  # L
            m_val = get_cell_value_with_merge(ws, row, 13)  # M

            c_norm2 = normalize_text(c_val)
            c_norm2 = re.sub(r"[（）\(\)【】\[\]{}]", "", c_norm2)

            if "小計" in c_norm2:
                continue

            if normalize_text(l_val) == "" and m_val is not None:
                ws.cell(row=row, column=12).value = m_val

        # ③ EFG列を削除
        ws.delete_cols(5, 3)

        # ④ さらにE列を削除
        ws.delete_cols(5)

    # ★ 選択したシート以降を、先頭シートに統合
    consolidate_selected_sheets(wb, target_sheet_names, add_blank_row=True)

    output_path = xlsx_path.replace(".xlsx", "_office_multi_converted.xlsx")
    wb.save(output_path)
    wb.close()
    return output_path
def get_cell_value_with_merge(ws, row, col):
    """
    結合セル対応：
    指定セルが結合範囲内なら、左上セルの値を返す
    """
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value

    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(
                row=merged_range.min_row,
                column=merged_range.min_col
            ).value
    return None


def normalize_text(v):
    """半角・全角スペースを除去（Noneは空扱い）"""
    if v is None:
        return ""
    if not isinstance(v, str):
        v = str(v)
    return re.sub(r"[\s\u3000]+", "", v)



# --- 各社変換処理 ---
def office_int(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows_to_delete = [idx for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2)
                      if row[1] is None or not isinstance(row[1], (int, float))]
    for idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(idx)

    ws.delete_cols(17)
    ws.delete_cols(15)
    ws.delete_cols(13)
    ws.delete_cols(8, 2)
    ws.delete_cols(4, 3)
    ws.insert_cols(4)

    pattern_total = re.compile(r'\b小\s*計\b')
    pattern_mid_total = re.compile(r'\b中\s*計\b')  # ★ 追加
    pattern_discount = re.compile(r'\b値\s*引\b')

    
    for row in ws.iter_rows(min_row=2):
        h_cell = row[7]
        c_cell = row[2]
    
        # ★ C列が完全に空白の行は処理しないで次へ進む
        if c_cell.value is None:
            continue

        c_text = str(c_cell.value)

        # ★ 中計 も 小計 と同様に除外する
        if (
            (h_cell.value is None)
            and (not pattern_total.search(c_text))
            and (not pattern_mid_total.search(c_text))   # ← ここ追加
            and (not pattern_discount.search(c_text))
        ):
            h_cell.value = row[8].value
    for row in ws.iter_rows(min_row=2):
        for cell in row[2:5]:
            if cell.value is not None and isinstance(cell.value, str):
                cell.value = cell.value.strip()

    output_path = xlsx_path.replace(".xlsx", "_converted.xlsx")
    wb.save(output_path)
    wb.close()
    return output_path

def comany(xlsx_path):
    # もし拡張子が .xls なら .xlsx に変換
    if xlsx_path.lower().endswith(".xls"):
        xlsx_converted = xlsx_path + "x"
        save_book_as(file_name=xlsx_path, dest_file_name=xlsx_converted)
        xlsx_path = xlsx_converted

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "内訳書" not in wb.sheetnames:
        wb.close()
        return None
    ws = wb["内訳書"]
    max_row = ws.max_row

    col = dict(b=2, c=3, f=6, g=7, i=9, j=10, k=11, l=12, m=13, n=14, o=15)

    # --- W/D/HをまとめてI列に出力 ---
    def fmt_dim(v):
        if v is None:
            return None
        try:
            f = float(v)
            return str(int(f)) if f.is_integer() else str(f)
        except:
            return str(v)

    def nonzero(v):
        if v in (None, "", 0, 0.0):
            return False
        try:
            return float(v) != 0
        except:
            return True

    for r in range(2, max_row + 1):
        w = ws.cell(r, col["i"]).value
        d = ws.cell(r, col["j"]).value
        h = ws.cell(r, col["k"]).value
        if w or d or h:
            merged = ""
            if nonzero(w): merged += f"W{fmt_dim(w)}"
            if nonzero(d): merged += f"D{fmt_dim(d)}"
            if nonzero(h): merged += f"H{fmt_dim(h)}"
            ws.cell(r, col["i"]).value = merged
            ws.merge_cells(start_row=r, start_column=col["i"], end_row=r, end_column=col["k"])
            ws.cell(r, col["i"]).alignment = Alignment(horizontal="left", vertical="center")

    # --- 【】内の文字をF列に抽出 ---
    pattern = re.compile(r"【(.*?)】")
    for r in range(2, max_row + 1):
        g_val = ws.cell(r, col["g"]).value
        g_str = str(g_val) if g_val else ""
        match = pattern.search(g_str)
        if match:
            ws.cell(r, col["f"]).value = match.group(1)

        if "施工" in g_str:
            l_val = ws.cell(r, col["l"]).value
            m_val = ws.cell(r, col["m"]).value
            try:
                merged = f"{float(l_val):.1f}{m_val}"
            except:
                merged = f"{l_val}{m_val}"
            ws.cell(r, col["i"]).value = merged
            ws.cell(r, col["l"]).value = 1
            ws.cell(r, col["m"]).value = "式"

    # --- ■や◆のある行をF列に転記 ---
    for r in range(2, max_row + 1):
        b_val = ws.cell(r, col["b"]).value
        c_val = ws.cell(r, col["c"]).value
        if b_val and "■" in str(b_val):
            ws.cell(r, col["f"]).value = mojimoji.han_to_zen(str(b_val))
        if c_val and "◆" in str(c_val):
            ws.cell(r, col["f"]).value = mojimoji.han_to_zen(str(c_val))

    # --- N列が空でO列が0以外 → NにOをコピー ---
    for r in range(2, max_row + 1):
        n_val = ws.cell(r, col["n"]).value
        o_val = ws.cell(r, col["o"]).value
        if (n_val is None or str(n_val).strip() == "") and nonzero(o_val):
            ws.cell(r, col["n"]).value = o_val

    # --- 列削除と挿入（元コードと同じ順序）---
    ws.delete_cols(10, 2)
    ws.delete_cols(7, 2)
    ws.delete_cols(1, 5)
    ws.insert_cols(2)

    # ★★★ A列が「部材」の行を特別処理 ★★★
    for r in range(2, ws.max_row + 1):
        a_val = ws.cell(r, 1).value  # A列（列削除後）
        if isinstance(a_val, str) and a_val.strip() == "部材":
            ws.cell(r, 1).value = "部材合計"
            for c in range(4, 8):  # D〜G
                ws.cell(r, c).value = None
    # ★★★ 完全な空行を削除（A〜H すべて空なら削除） ★★★
    for r in range(ws.max_row, 1, -1):  # 逆順で消す
        if all(ws.cell(r, c).value in (None, "") for c in range(1, 9)):  # A〜H列をチェック
            ws.delete_rows(r)


    # --- 不要シート削除（内訳書だけ残す）---
    for name in list(wb.sheetnames):
        if name != "内訳書":
            wb.remove(wb[name])

    output_path = xlsx_path.replace(".xlsx", "_comany_converted.xlsx")
    wb.save(output_path)
    wb.close()
    return output_path


# def soken(xlsx_path):
#     wb = openpyxl.load_workbook(xlsx_path, data_only=True)
#     ws = wb.active
#     for row in range(ws.max_row, 0, -1):
#         if all(ws.cell(row=row, column=col).value is None for col in range(1, ws.max_column + 1)):
#             ws.delete_rows(row)
#     ws.delete_rows(1, 8)
#     ws['N1'] = '単　価'

#     rows_to_delete = [idx for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2)
#                       if row[2] == '名　　　称']
#     for idx in sorted(rows_to_delete, reverse=True):
#         ws.delete_rows(idx)

#     pattern = re.compile(r"頁")
#     for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
#         if pattern.search(str(row[17])):
#             ws.delete_rows(idx)

#     output_path = xlsx_path.replace(".xlsx", "_converted.xlsx")
#     wb.save(output_path)
#     wb.close()
#     return output_path

# --- UI ---

# ▼ 既存UI
company = st.sidebar.selectbox(
    "仕入先を選択してください",
    ["オフィス（１シート）", "オフィス（複数シート）", "コマニー", "帝国倉庫PDF"]
)
# 仕入先変更を検知
prev_company = st.session_state.get("prev_company")

if prev_company is None:
    st.session_state["prev_company"] = company
elif prev_company != company:
    # ★ 仕入先が変わったら、仕入先ファイル uploader だけ初期化
    st.session_state["uploader_reset_counter"] = st.session_state.get("uploader_reset_counter", 0) + 1

    # ★ 仕入先側の状態だけ消す
    for k in [
        "office_multi_start_sheet",
        "src_cell",
        "dest_cell",
        "preview_sheet_name",
        "current_supplier_name",
    ]:
        if k in st.session_state:
            del st.session_state[k]

    # ★ 処理済み見積の表示用もクリア
    converted_excel_path = None
    selected_start_sheet = None
    df_converted = None

    # ★ 新しい仕入先を記録
    st.session_state["prev_company"] = company
    st.rerun()

uploader_reset_counter = st.session_state.get("uploader_reset_counter", 0)

supplier_file_type = ["pdf"] if company == "帝国倉庫PDF" else ["xls", "xlsx"]
supplier_file_label = "仕入先PDFファイルをアップロードしてください" if company == "帝国倉庫PDF" else "仕入先ファイル（xls/.xlsx）をアップロードしてください"

supplier_file = st.file_uploader(
    supplier_file_label,
    type=supplier_file_type,
    key=f"supplier_file_{uploader_reset_counter}"
)

converted_excel_path = None
selected_start_sheet = None
df_converted = None
preview_sheet_name = None

# ★ ここで先に処理開始シートを表示
if supplier_file:
    ext = os.path.splitext(supplier_file.name)[1].lower()
    supplier_file_bytes = supplier_file.getvalue()
    raw_path = save_uploaded_bytes(supplier_file_bytes, suffix=ext)
    if company == "帝国倉庫PDF":
        df_converted = load_teisoh_detail_df(raw_path)
        converted_excel_path = os.path.join(APP_TEMP_DIR, "teisoh_pdf_converted.xlsx")
        with pd.ExcelWriter(converted_excel_path, engine="openpyxl") as writer:
            df_converted.to_excel(writer, sheet_name="帝国倉庫PDF", index=False)
        preview_sheet_name = "帝国倉庫PDF"
    else:
        xlsx_path = convert_xls_to_xlsx(raw_path) if ext == ".xls" else raw_path

        if company == "オフィス（複数シート）":
            raw_wb = load_workbook(xlsx_path, data_only=True)
            raw_sheet_options = raw_wb.sheetnames
            raw_wb.close()

            selected_start_sheet = st.selectbox(
                "処理開始シートを選択してください（このシート以降を処理）",
                raw_sheet_options,
                key="office_multi_start_sheet"
            )

# ★ selectbox のあとにテンプレート uploader
template_file = st.file_uploader(
    "見積テンプレートExcel（.xlsx）をアップロード",
    type=["xlsx"],
    key=f"template_file_{uploader_reset_counter}"
)

# ★ 変換処理
if supplier_file:
    if company == "オフィス（１シート）":
        converted_excel_path = office_int(xlsx_path)

    elif company == "オフィス（複数シート）":
        converted_excel_path = office_int_multi(xlsx_path, selected_start_sheet)

    elif company == "コマニー":
        converted_excel_path = comany(xlsx_path)

    if converted_excel_path is None:
        if company == "コマニー":
            st.error("この見積ファイルには『内訳書』シートが無いため、コマニーの変換ができませんでした。")
        else:
            st.error(f"{company} の変換に失敗しました。（対応シートが無い/形式が違う可能性があります）")
        st.stop()

    wb = load_workbook(converted_excel_path)

    if company == "オフィス（複数シート）" and selected_start_sheet:
        preview_sheet_name = selected_start_sheet
    else:
        preview_sheet_name = wb.sheetnames[0]

    ws = wb[preview_sheet_name]
    data = list(ws.values)
    df_converted = pd.DataFrame(data)
    df_converted.columns = [get_column_letter(i + 1) for i in range(len(df_converted.columns))]
    df_converted.index += 1
    df_converted = df_converted.fillna("")



# ★「初回だけ」テンプレを確定させる（2回目以降の rerun では上書きしない）
# テンプレは「新規モード」では毎回これを基準にしたいので、アップロードがあれば更新してOK
if template_file:
    template_bytes = template_file.getvalue()
    template_path = save_uploaded_bytes(template_bytes, suffix=".xlsx")
    st.session_state.original_template_path = template_path

    if not st.session_state.get("result_path"):
        st.session_state.template_path = template_path
        st.session_state.template_wb = load_workbook(template_path)


# テンプレート未読込 → 見積だけ単体表示
if df_converted is not None and not st.session_state.template_wb:
    st.subheader("処理済み見積ファイルのプレビュー")
    st.dataframe(df_converted, use_container_width=True)

# テンプレートが読み込まれている場合 → 両方表示
elif df_converted is not None and st.session_state.template_wb:
    sheet_name = st.selectbox("見積テンプレートシートを選んでください", st.session_state.template_wb.sheetnames, key="sheet_selector")
    ws_template = st.session_state.template_wb[sheet_name]
    template_data = list(ws_template.values)
    df_template = pd.DataFrame(template_data).fillna("")
    df_template.columns = [get_column_letter(i+1) for i in range(len(df_template.columns))]
    df_template.index += 1

    col1, col2 = st.columns(2)
    with col1:
        st.subheader(f"処理済み見積ファイルのプレビュー（{preview_sheet_name}）")
        st.dataframe(df_converted, use_container_width=True)
    with col2:
        # シート別の転記結果があればそれを表示
        if (
            "updated_template_df_map" in st.session_state
            and sheet_name in st.session_state.updated_template_df_map
        ):
            st.subheader(f"見積テンプレートシート：転記結果（{sheet_name}）")
            st.data_editor(
                st.session_state.updated_template_df_map[sheet_name],
                use_container_width=True,
                disabled=True,
                hide_index=False
            )
        else:
            st.subheader(f"見積テンプレートシート: {sheet_name}")
            st.data_editor(
                df_template,
                use_container_width=True,
                disabled=True,
                hide_index=False
            )




    # --- ★ ここから転記UI ---
    st.markdown("### 転記設定（開始セルと転記先セルを入力）")

    src = st.text_input("転記元の開始セル（例: C14）", key="src_cell")
    dest = st.text_input("転記先セル（例: B3）", key="dest_cell")
    output_name = st.text_input("出力ファイル名（拡張子不要）", value="merged_result", key="output_name")


    if st.button("✅ 一括転記する"):
        # ① 入力チェック（ロック前）
        src  = (st.session_state.get("src_cell")  or "").strip().upper()
        dest = (st.session_state.get("dest_cell") or "").strip().upper()

        if not src:
            st.error("転記元の開始セル（例: C2）を入力してください")
            st.stop()
        if not dest:
            st.error("転記先セル（例: B2）を入力してください")
            st.stop()

        # ② ロック獲得
        if not acquire_lock():
            st.warning("現在、他のユーザーが転記中です。少し待ってください。")
            st.stop()

        try:
            src_row = get_row_from_cell(src)
            df_to_write = df_converted.loc[src_row:].copy()

            cols_check = [c for c in ["A","B","C","D","E","F"] if c in df_to_write.columns]
            if cols_check:
                df_to_write = df_to_write[
                    ~df_to_write[cols_check]
                    .astype(str)
                    .apply(lambda x: x.str.strip())
                    .eq("")
                    .all(axis=1)
                ]

            # ここで必ず「今画面に入っている値」を取り直す
            output_name = (st.session_state.get("output_name") or "merged_result").strip()

            # ファイル名に使えない文字を除去（念のため）
            output_name = re.sub(r'[\\/:*?"<>|]', '_', output_name)

            ts = datetime.datetime.now().strftime("%m%d")

            # ===== 差し替え形式：継続でもファイル名を更新しつつ、最終的に1ファイルにする =====
            last_result_path = st.session_state.get("last_result_path")
            original_template_path = st.session_state.get("original_template_path")

            # 入力ファイル名（拡張子つき）
            new_filename = f"{ts}_{output_name}.xlsx"
            new_path = os.path.join(RESULT_DIR, new_filename)

            if last_result_path:
                # 継続：基本は上書き
                base_template_path = last_result_path
                result_path = last_result_path

                # ただし、ユーザーが別名を入れたら「その名前に切り替える」
                if os.path.basename(last_result_path) != new_filename:
                    # 既に同名があるなら消しておく（上書き許可）
                    if os.path.exists(new_path):
                        os.remove(new_path)

                    shutil.copy(last_result_path, new_path)
                    base_template_path = new_path
                    result_path = new_path

            else:
                # 1回目（新規）
                base_template_path = original_template_path
                if not base_template_path:
                    raise RuntimeError("テンプレートが未設定です。先にテンプレートExcelをアップロードしてください。")
                result_path = new_path

            # ======================================================================


            # ★★★ ここまで ★★★

            write_df_to_template_com(
                template_path=base_template_path,
                out_path=result_path,
                sheet_name=sheet_name,
                start_cell=dest,
                df_src=df_to_write,
                col_map=get_col_map(company),
            )
            # --- 差し替え：前回ファイル名と今回のファイル名が違うなら、前回ファイルを削除して1本化 ---
            if last_result_path and os.path.abspath(last_result_path) != os.path.abspath(result_path):
                try:
                    os.remove(last_result_path)
                except Exception:
                    pass


            st.session_state["download_name"] = os.path.basename(result_path)
            st.session_state["last_result_path"] = result_path  # 次回継続の基準
            st.session_state["result_path"] = result_path       # 画面表示/ダウンロード用
            st.session_state.template_path = result_path
            st.session_state.template_wb = load_workbook(result_path)


            st.success("転記が完了しました")
            release_lock()   # ✅ rerun前に解除（保険）
            st.rerun()

        except Exception as e:
            st.error(str(e))
            release_lock()
            st.stop()

        finally:
            release_lock()




# ★ 一括転記後だけ、ダウンロードボタンを表示
if "result_path" in st.session_state:
    # いま画面に入っている出力名を「DL名」に反映（転記後に変えてもOK）
    dl_base = (st.session_state.get("output_name") or "merged_result").strip()
    dl_base = re.sub(r'[\\/:*?"<>|]', '_', dl_base)  # 禁止文字除去

    # ts をファイル名に付けたいなら：転記時に使った ts を保存しておくのが理想
    # いったん簡易に「保存先ファイルの先頭4桁(0109_)を流用」する例：
    current_base = os.path.basename(st.session_state["result_path"])
    prefix = current_base.split("_", 1)[0]  # 0109 だけ取り出す（無い場合もある）

    # prefix が4桁じゃなければ付けない
    if prefix.isdigit() and len(prefix) == 4:
        dl_name = f"{prefix}_{dl_base}.xlsx"
    else:
        dl_name = f"{dl_base}.xlsx"

    with open(st.session_state["result_path"], "rb") as f:
        st.download_button(
            "転記済みファイルをダウンロード",
            f,
            file_name=dl_name,
            key="download_btn"  # key固定（表示が安定）
        )

    st.write("保存先:", st.session_state.get("result_path"))
    st.write("DL名:", dl_name)






                # --- ★ ここまで転記UI ---
